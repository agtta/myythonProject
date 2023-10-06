import openpyxl

workbook = openpyxl.load_workbook('sales.xlsx')
worksheet = workbook.active
lista = []

print(worksheet)
for i in worksheet:
    print(i)

for i in range(0, worksheet.max_row):
    for col in worksheet.iter_cols(1, worksheet.max_column):
            lista.append(col[i].value)

print(lista)

